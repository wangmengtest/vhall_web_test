#!/usr/bin/env python
# encoding:utf8
import require
from openpyxl import Workbook

"""
公共Excel写文件处理 单例
"""


class ExcelWriteUtil(object):
    filePath = None
    workSpace = None
    workSheet = None

    def __init__(self, filePath):
        print(filePath)
        self.filePath = filePath
        self.workSpace = Workbook()

        # 默认激活几个sheet
        self.workSheet = self.workSpace.active

    # 设置工作表明
    def createSheet(self, sheetName):
        self.workSheet = self.workSpace.create_sheet(sheetName, 0)

    def setTitle(self, tableList):
        # 维护表头
        #        if row < 1 or column < 1:
        #          raise ValueError("Row or column values must be at least 1")
        # 如上，openpyxl 的首行、首列 是 （1,1）而不是（0,0），如果坐标输入含有小于1的值，提示 ：Row or column values must be at least 1，即最小值为1.
        for col in range(len(tableList)):
            c = col + 1
            self.workSheet.cell(row=1, column=c).value = tableList[col]

    # 通过列表增加数据 [[1,2,3,4], [2,2,3,4]]
    def appendByList(self, listData):
        if not isinstance(listData[0], list):
            listData = [listData]

        for row in range(len(listData)):
            self.workSheet.append(listData[row])

    # 通过字典新增数据 {'A' : 'xx', 'B','xxx}
    def appendByDict(self, dictData):
        self.workSheet.append(dictData)

    def save(self):
        self.workSpace.save(filename=self.filePath)



