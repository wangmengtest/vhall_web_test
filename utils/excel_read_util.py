#!/usr/bin/env python
# encoding:utf8
import os
import openpyxl
import require
from utils.log_util import logUtil
from utils.string_util import stringUtil

"""
公共Excel读文件处理 单例
"""


class ExcelReadUtil(object):
    filePath = None
    workSpace = None
    workSheet = None

    def __init__(self, filePath):
        self.filePath = filePath
        if not os.path.exists(self.filePath):
            logUtil.renderConsole(self.filePath + " 文件不存在")
            exit(0)
        self.workSpace = openpyxl.load_workbook(filePath)

        self.setSheet()

    # 设置 工作表
    def setSheet(self, sheetName=None):
        if sheetName:
            if sheetName in self.workSpace:
                self.workSheet = self.workSpace[sheetName]
            else:
                logUtil.error(sheetName + " 表不存在")
                exit(0)
        else:
            self.workSheet = self.workSpace.active


    # 获取最大行
    def getHeight(self):
        return self.workSheet.max_row

    # 获取最大列
    def getWidth(self):
        return self.workSheet.max_column

    # 按给定的索引将表格映射进去
    def mapToIndex(self, startNum, indexArray):
        result = []
        for i in range(startNum, self.getHeight() + 1):
            rowList = {}
            allValueNull = True

            for index, value in enumerate(indexArray):
                column = stringUtil.getBigAbcFromIndex(index)
                columnValue = self.getValue(column, i)
                if columnValue:
                    allValueNull = False
                rowList[value] = str(columnValue).strip().encode().decode("utf-8-sig")

            # 所有值都为空时，说明返回高度不准确，不添加
            if not allValueNull:
                result.append(rowList)

        return result

    # 获取值
    def getValue(self, column, row):
        index = str(column) + str(row)

        return self.workSheet[index].value